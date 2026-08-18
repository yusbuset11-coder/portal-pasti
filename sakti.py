import streamlit as st
import google.generativeai as genai
import pandas as pd
import io
import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def render_sakti():
    st.markdown("### 🎯 SAKTI (Sistem Asesmen & Kompetensi Terintegrasi)")
    st.write("Sistem Asesmen & Kompetensi Terintegrasi (Pembelajaran Mendalam)")

    # Konfigurasi Link CSV Google Spreadsheet Database_PASTI_Pusat
    with st.expander("🔗 Konfigurasi Sumber Data Siswa (Google Sheets)", expanded=False):
        default_csv = st.session_state.get("sakti_csv_url", "https://docs.google.com/spreadsheets/d/1terQdXNZX1aESF0G02uSn9R7eKLKDgbkit11GpX1pA/export?format=csv&sheet=Siswa")
        csv_url = st.text_input("Link CSV Google Sheet (Sheet: Siswa)", value=default_csv)
        if st.button("Simpan Link Database"):
            st.session_state["sakti_csv_url"] = csv_url
            st.success("Link database siswa berhasil diperbarui!")

    # Fungsi untuk memuat data siswa langsung dari Google Sheet dengan Fallback Data Lengkap
    @st.cache_data(ttl=60)
    def load_data_siswa_from_sheet():
        url = st.session_state.get("sakti_csv_url", "https://docs.google.com/spreadsheets/d/1terQdXNZX1aESF0G02uSn9R7eKLKDgbkit11GpX1pA/export?format=csv&sheet=Siswa")
        try:
            df_siswa = pd.read_csv(url)
            df_siswa.columns = df_siswa.columns.str.strip()
            if not df_siswa.empty and "ID_Siswa" in df_siswa.columns:
                return df_siswa
        except Exception:
            pass
        
        # Fallback Data Sesuai Sampel Database PASTI Pusat
        fallback_data = [
            {"ID_Siswa": 1, "Sekolah": "SMK Negeri 2 Bangkalan", "Kelas": "X-1", "Nama_Siswa": "AISYAH KHOIRUMNISA"},
            {"ID_Siswa": 2, "Sekolah": "SMK Negeri 2 Bangkalan", "Kelas": "X-1", "Nama_Siswa": "ALDO PUJI FEBRIANSYAH"},
            {"ID_Siswa": 3, "Sekolah": "SMK Negeri 2 Bangkalan", "Kelas": "X-1", "Nama_Siswa": "ASSYA RAHMADHANA"},
            {"ID_Siswa": 4, "Sekolah": "SMK Negeri 2 Bangkalan", "Kelas": "X-1", "Nama_Siswa": "AUFAR NURIEL ADLI RIFA'I"},
            {"ID_Siswa": 5, "Sekolah": "SMK Negeri 2 Bangkalan", "Kelas": "X-1", "Nama_Siswa": "AULYA ZIVANA LETISYA"},
            {"ID_Siswa": 6, "Sekolah": "SMK Negeri 2 Bangkalan", "Kelas": "X-1", "Nama_Siswa": "AVRILLIA ZIBA AQILLA NUR AULIYAH"},
            {"ID_Siswa": 7, "Sekolah": "SMK Negeri 2 Bangkalan", "Kelas": "X-1", "Nama_Siswa": "AZKA RANENDRA"},
            {"ID_Siswa": 8, "Sekolah": "SMK Negeri 2 Bangkalan", "Kelas": "X-1", "Nama_Siswa": "BEBY JOSEPHIRA FLORIDIA"},
            {"ID_Siswa": 9, "Sekolah": "SMK Negeri 2 Bangkalan", "Kelas": "X-1", "Nama_Siswa": "FABIAN NAROTAMA ATILA SETIA"},
            {"ID_Siswa": 10, "Sekolah": "SMK Negeri 2 Bangkalan", "Kelas": "X-1", "Nama_Siswa": "FAUZIA NUR ASKIYAH"},
            {"ID_Siswa": 1, "Sekolah": "SMK Negeri 2 Bangkalan", "Kelas": "X-2", "Nama_Siswa": "HAVIDZAH SITI AZZAHRA"},
            {"ID_Siswa": 2, "Sekolah": "SMK Negeri 2 Bangkalan", "Kelas": "X-2", "Nama_Siswa": "HEBRY ADITYA JOVAN ELNANDO"},
            {"ID_Siswa": 3, "Sekolah": "SMK Negeri 2 Bangkalan", "Kelas": "X-2", "Nama_Siswa": "IBRAZIEL WIZARD ERDHANA"},
            {"ID_Siswa": 4, "Sekolah": "SMK Negeri 2 Bangkalan", "Kelas": "X-2", "Nama_Siswa": "KAYLA MARCHILIA AZZAHRA"},
            {"ID_Siswa": 5, "Sekolah": "SMK Negeri 2 Bangkalan", "Kelas": "X-2", "Nama_Siswa": "KENNITIIJA DIVANA PARAWANZA"},
            {"ID_Siswa": 6, "Sekolah": "SMK Negeri 2 Bangkalan", "Kelas": "X-2", "Nama_Siswa": "KHAYLA AZHARA NURSUCAHYO"},
            {"ID_Siswa": 7, "Sekolah": "SMK Negeri 2 Bangkalan", "Kelas": "X-2", "Nama_Siswa": "MOCH. FADHIL APRILLIANSYAH"},
            {"ID_Siswa": 8, "Sekolah": "SMK Negeri 2 Bangkalan", "Kelas": "X-2", "Nama_Siswa": "MUHAMMAD ALIF"},
            {"ID_Siswa": 9, "Sekolah": "SMK Negeri 2 Bangkalan", "Kelas": "X-2", "Nama_Siswa": "NADHİFA SHAFA SALSABİLLA"},
            {"ID_Siswa": 10, "Sekolah": "SMK Negeri 2 Bangkalan", "Kelas": "X-2", "Nama_Siswa": "NOVITA PUTRI ZALSABİLLA"}
        ]
        return pd.DataFrame(fallback_data)

    df_master_siswa = load_data_siswa_from_sheet()

    # Form Pembuatan Soal & Asesmen (AI)
    with st.expander("✨ Parameter Pembuatan Soal & Asesmen (Pendekatan PM)", expanded=False):
        col1, col2 = st.columns(2)
        with col1:
            mapel_ai = st.text_input("Mata Pelajaran (AI)", placeholder="Contoh: Bahasa Indonesia")
        with col2:
            materi_ai = st.text_input("Materi / Topik (AI)", placeholder="Contoh: Mengidentifikasi Makna Kata")

        col3, col4, col5 = st.columns(3)
        with col3:
            jenjang_ai = st.selectbox("Jenjang (AI)", ["SD", "SMP", "SMA", "SMK"], index=3)
        with col4:
            fase_ai = st.selectbox("Fase (AI)", ["Fase A", "Fase B", "Fase C", "Fase D", "Fase E", "Fase F"], index=4)
        with col5:
            kelas_ai = st.selectbox("Kelas (AI)", [f"Kelas {i}" for i in range(1, 13)], index=9)

        col6, col7 = st.columns(2)
        with col6:
            jenis_asesmen_ai = st.selectbox("Jenis Asesmen (AI)", ["Asesmen Formatif", "Asesmen Sumatif"])
        with col7:
            if jenis_asesmen_ai == "Asesmen Formatif":
                sub_asesmen_ai = st.selectbox("Sub Jenis Asesmen (AI)", ["Tertulis", "Tak Tertulis"])
            else:
                sub_asesmen_ai = st.selectbox("Sub Jenis Asesmen (AI)", ["Tulis", "Lisan", "Tugas", "Praktik", "Proyek", "Produk"])

        col8, col9 = st.columns(2)
        with col8:
            jumlah_soal = st.number_input("Jumlah Butir Soal", min_value=1, max_value=20, value=5)
        with col9:
            kesulitan = st.selectbox("Tingkat Kesulitan", ["Mudah", "Sedang", "Sulit"], index=1)

        btn_generate = st.button("✨ Buat Instrumen Asesmen dengan Gemini AI 🚀", use_container_width=True)

    if btn_generate:
        if not mapel_ai or not materi_ai:
            st.warning("Harap isi Mata Pelajaran dan Materi terlebih dahulu!")
        else:
            with st.spinner("Sedang merancang instrumen asesmen mendalam..."):
                aturan_opsi = "3 opsi (A sampai C)" if jenjang_ai == "SD" else ("4 opsi (A sampai D)" if jenjang_ai == "SMP" else "5 opsi (A sampai E)")
                prompt = f"""
                Buatkan instrumen asesmen dengan pendekatan Pembelajaran Mendalam (PM).
                Mata Pelajaran: {mapel_ai}, Materi: {materi_ai}, Jenjang: {jenjang_ai}, Fase: {fase_ai}, Kelas: {kelas_ai}, 
                Jenis Asesmen: {jenis_asesmen_ai}, Sub Jenis: {sub_asesmen_ai}, Jumlah Soal: {jumlah_soal}, Tingkat Kesulitan: {kesulitan}, Aturan Pilihan Ganda: {aturan_opsi}
                """
                try:
                    api_key = st.session_state.get("gemini_api_key", "")
                    if not api_key:
                        st.error("API Key belum diset. Silakan masukkan di sidebar.")
                    else:
                        genai.configure(api_key=api_key)
                        model = genai.GenerativeModel("gemini-1.5-flash")
                        response = model.generate_content(prompt)
                        st.markdown("### 📋 Hasil Instrumen Asesmen dari AI:")
                        st.success("Instrumen berhasil dibuat!")
                        st.markdown(response.text)
                except Exception as e:
                    st.error(f"Gagal memanggil AI: {str(e)}")

    st.markdown("---")

    st.markdown("### 📄 Lembar Input Hasil Asesmen Siswa")
    st.write("Catat dan rekap nilai hasil asesmen formatif atau sumatif peserta didik.")

    list_opsi_sekolah = df_master_siswa["Sekolah"].dropna().unique().tolist() if not df_master_siswa.empty else ["SMK Negeri 2 Bangkalan"]

    with st.form("form_input_asesmen_sakti"):
        col_a, col_b = st.columns(2)
        with col_a:
            tanggal_input = st.date_input("Tanggal", value=datetime.date.today())
            nama_guru = st.text_input("Nama Guru", placeholder="Masukkan nama guru pengampu")
            mata_pelajaran = st.text_input("Mata Pelajaran", placeholder="Masukkan mata pelajaran")
            nama_sekolah = st.selectbox("Pilih Asal Sekolah", list_opsi_sekolah)
            jenjang_kelas = st.selectbox("Pilih Jenjang", ["SD", "SMP", "SMA", "SMK"], index=3)
            
        with col_b:
            if not df_master_siswa.empty:
                list_opsi_kelas = df_master_siswa[df_master_siswa["Sekolah"] == nama_sekolah]["Kelas"].dropna().unique().tolist()
            else:
                list_opsi_kelas = ["X-1", "X-2"]
            
            pilih_kelas = st.selectbox("Pilih Kelas", list_opsi_kelas if list_opsi_kelas else ["X-1"])
            
            jenis_asesmen = st.selectbox("Pilih Jenis Asesmen", ["Formatif", "Sumatif"])
            
            if jenis_asesmen == "Formatif":
                sub_jenis_asesmen = st.selectbox("Subjenis Asesmen", ["Tertulis", "Tak Tertulis"])
            else:
                sub_jenis_asesmen = st.selectbox("Subjenis Asesmen", ["Tulis", "Lisan", "Tugas", "Praktik", "Proyek", "Produk"])
                
            materi_topik = st.text_input("Ketik Materi", placeholder="Masukkan materi pembelajaran")

            # Filter ID_Siswa / Nomor Absen berdasarkan Sekolah & Kelas
            if not df_master_siswa.empty:
                df_filtered_siswa = df_master_siswa[(df_master_siswa["Sekolah"] == nama_sekolah) & (df_master_siswa["Kelas"] == pilih_kelas)]
                list_id_siswa = sorted(df_filtered_siswa["ID_Siswa"].dropna().unique().tolist())
            else:
                df_filtered_siswa = pd.DataFrame()
                list_id_siswa = [1, 2, 3]

            no_absen = st.selectbox("Pilih Nomor Absen (ID_Siswa)", list_id_siswa if list_id_siswa else [1])
            
            # Otomatis ambil Nama Siswa
            nama_siswa_otomatis = ""
            if not df_filtered_siswa.empty:
                match_row = df_filtered_siswa[df_filtered_siswa["ID_Siswa"] == no_absen]
                if not match_row.empty:
                    nama_siswa_otomatis = str(match_row.iloc[0]["Nama_Siswa"])

            nama_siswa = st.text_input("Nama Siswa (Otomatis dari Sheet Siswa)", value=nama_siswa_otomatis)

        # Input Nilai
        nilai_siswa = st.number_input("Ketik Nilai Asesmen (Skala 0 - 100)", min_value=0.0, max_value=100.0, value=75.0, step=0.25)
        
        # Logika Catatan Otomatis Berdasarkan Nilai
        if nilai_siswa < 60:
            default_catatan = "Perlu Bimbingan!"
        elif 60 <= nilai_siswa <= 70:
            default_catatan = "Belajarlah lebih giat!"
        elif 71 <= nilai_siswa <= 80:
            default_catatan = "Tingkatkan prestasi belajar Anda!"
        elif 81 <= nilai_siswa <= 90:
            default_catatan = "Pertahankan prestasi belajar Anda!"
        else:
            default_catatan = "Luar biasa! Pertahankan prestasi belajar Anda!"

        catatan_guru = st.text_area("Catatan Perkembangan / Umpan Balik (Feedback)", value=default_catatan)

        submitted = st.form_submit_button("💾 Simpan Nilai Asesmen")

        if submitted:
            if not nama_siswa.strip() or not nama_guru.strip() or not nama_sekolah.strip():
                st.error("❌ Asal Sekolah, Nama Guru, dan Nama Siswa wajib diisi sebelum menyimpan data!")
            else:
                if "sakti_data_nilai" not in st.session_state:
                    st.session_state["sakti_data_nilai"] = []
                
                st.session_state["sakti_data_nilai"].append({
                    "Sekolah": nama_sekolah,
                    "Tanggal": str(tanggal_input),
                    "Guru Pengampu": nama_guru,
                    "Mata Pelajaran": mata_pelajaran,
                    "Jenjang": jenjang_kelas,
                    "Kelas": pilih_kelas,
                    "Jenis Asesmen": jenis_asesmen,
                    "Subjenis Asesmen": sub_jenis_asesmen,
                    "Materi": materi_topik,
                    "No Absen": no_absen,
                    "Nama Siswa": nama_siswa,
                    "Nilai": nilai_siswa,
                    "Catatan": catatan_guru if catatan_guru else "-"
                })
                st.success(f"✅ Hasil asesmen untuk **{nama_siswa}** (No Absen {no_absen}) berhasil disimpan!")

    # Fungsi untuk membuat file Excel yang rapi dengan openpyxl
    def generate_styled_excel(df):
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Rekap Asesmen')
            workbook = writer.book
            worksheet = writer.sheets['Rekap Asesmen']
            
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            align_center = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
            )
            
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = align_center
            
            for row in range(2, len(df) + 2):
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center")
                    
            for col in worksheet.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                worksheet.column_dimensions[col_letter].width = max(max_len + 4, 15)
                
        output.seek(0)
        return output.getvalue()

    # Menampilkan Tabel Rekap Nilai & Filter
    if "sakti_data_nilai" in st.session_state and st.session_state["sakti_data_nilai"]:
        st.markdown("---")
        st.subheader("📊 Tabel Rekapitulasi Hasil Asesmen Kelas")
        
        df_all = pd.DataFrame(st.session_state["sakti_data_nilai"])

        st.markdown("#### 🔍 Filter Rekapitulasi Data")
        f_col1, f_col2, f_col3 = st.columns(3)
        
        list_sekolah_filter = ["Semua Sekolah"] + sorted(df_all["Sekolah"].unique().tolist())
        list_kelas_filter = ["Semua Kelas"] + sorted(df_all["Kelas"].unique().tolist())
        list_mapel_filter = ["Semua Mata Pelajaran"] + sorted(df_all["Mata Pelajaran"].unique().tolist())

        with f_col1:
            pilih_filter_sekolah = st.selectbox("Filter Sekolah", list_sekolah_filter)
        with f_col2:
            pilih_filter_kelas = st.selectbox("Filter Kelas", list_kelas_filter)
        with f_col3:
            pilih_filter_mapel = st.selectbox("Filter Mata Pelajaran", list_mapel_filter)

        df_filtered = df_all.copy()
        if pilih_filter_sekolah != "Semua Sekolah":
            df_filtered = df_filtered[df_filtered["Sekolah"] == pilih_filter_sekolah]
        if pilih_filter_kelas != "Semua Kelas":
            df_filtered = df_filtered[df_filtered["Kelas"] == pilih_filter_kelas]
        if pilih_filter_mapel != "Semua Mata Pelajaran":
            df_filtered = df_filtered[df_filtered["Mata Pelajaran"] == pilih_filter_mapel]

        st.dataframe(df_filtered, use_container_width=True)

        if not df_filtered.empty:
            excel_data = generate_styled_excel(df_filtered)
            st.download_button(
                label="📥 Unduh Rekap Nilai Terfiltrasi (Excel .xlsx)",
                data=excel_data,
                file_name="Rekapitulasi_Hasil_Asesmen_Sakti.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_rekap_excel"
            )
        else:
            st.info("ℹ️ Tidak ada data yang sesuai dengan filter yang dipilih.")
    else:
        st.info("ℹ️ Belum ada data nilai siswa yang diinput.")