import base64
from datetime import date
import io
import json
from google.oauth2.service_account import Credentials
import gspread
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import streamlit as st

# Default nama guru (tetap disimpan di database untuk rekam jejak)
NAMA_GURU = "Yustinus Budi Setyanta"


def get_google_sheet_connection():
  scope = [
      "https://www.googleapis.com/auth/spreadsheets",
      "https://www.googleapis.com/auth/drive",
  ]
  credentials_dict = json.loads(
      base64.b64decode(st.secrets["gcp_base64"]).decode("utf-8")
  )
  creds = Credentials.from_service_account_info(
      credentials_dict, scopes=scope
  )
  client = gspread.authorize(creds)
  sheet = client.open("Database_PASTI_Pusat")
  return sheet


def render_digma_module():
  st.markdown("### 📊 DIGMA: Digitalisasi Jurnal Mengajar")

  try:
    sh = get_google_sheet_connection()
    jurnal_ws = sh.worksheet("Jurnal_Mengajar")
    siswa_ws = sh.worksheet("Siswa")
    df_siswa = pd.DataFrame(siswa_ws.get_all_records())
    daftar_kelas = (
        df_siswa["Kelas"].dropna().unique().tolist()
        if "Kelas" in df_siswa.columns
        else ["Belum ada kelas"]
    )
  except Exception as e:
    st.error(f"Gagal memuat database: {e}")
    return

  tab1, tab2, tab3 = st.tabs(
      ["✍️ Input Jurnal", "📋 Daftar Jurnal", "📥 Export Laporan"]
  )

  # --- TAB 1: INPUT JURNAL ---
  with tab1:
    st.markdown("#### 📝 Form Input Jurnal Mengajar Harian")

    with st.form("form_jurnal_mengajar", clear_on_submit=True):
      st.markdown("##### 🏫 Lokasi & Jadwal Mengajar")
      col_s1, col_s2 = st.columns(2)
      with col_s1:
        sekolah_pilihan = st.text_input(
            "Nama Sekolah Tempat Mengajar",
            value="SMK Negeri 2 Bangkalan",
            placeholder="Contoh: SMK Negeri 1 Kwanyar",
        )
      with col_s2:
        tanggal = st.date_input("📅 Tanggal Mengajar", value=date.today())

      col1, col2, col3 = st.columns(3)
      with col1:
        kelas = st.selectbox("🏫 Kelas", options=daftar_kelas)
      with col2:
        jp_ke = st.text_input(
            "⏱️ Jam Pelajaran (JP Ke-)", placeholder="Contoh: 1-2"
        )
      with col3:
        mata_pelajaran = st.text_input(
            "📖 Mata Pelajaran", placeholder="Contoh: Pendidikan Pancasila"
        )

      st.markdown("##### 👥 Kehadiran Siswa")
      col_a, col_b = st.columns(2)
      with col_a:
        hadir = st.number_input(
            "✅ Siswa Hadir", min_value=0, value=30, step=1
        )
      with col_b:
        tidak_hadir = st.number_input(
            "❌ Siswa Tidak Hadir", min_value=0, value=0, step=1
        )

      st.markdown("##### 📝 Materi & Refleksi Pembelajaran")
      topik_materi = st.text_area(
          "💡 Topik / Materi Pokok",
          placeholder="Tuliskan topik atau tujuan pembelajaran hari ini...",
      )
      catatan = st.text_area(
          "📌 Catatan / Refleksi",
          placeholder="Catatan tambahan atau refleksi kegiatan kelas...",
      )

      st.markdown("")
      submitted = st.form_submit_button(
          "💾 Simpan Jurnal Mengajar", use_container_width=True
      )

      if submitted:
        if not sekolah_pilihan or not mata_pelajaran or not topik_materi:
          st.warning(
              "⚠️ Mohon lengkapi Nama Sekolah, Mata Pelajaran, dan"
              " Topik/Materi!"
          )
        else:
          try:
            row_data = [
                str(tanggal),
                sekolah_pilihan,
                NAMA_GURU,
                mata_pelajaran,
                kelas,
                jp_ke,
                topik_materi,
                int(hadir),
                int(tidak_hadir),
                catatan,
            ]
            jurnal_ws.append_row(row_data)
            st.success("✅ Jurnal mengajar berhasil disimpan ke Database Pusat!")
            st.balloons()
          except Exception as e:
            st.error(f"❌ Gagal menyimpan jurnal ke Database: {e}")

  # --- TAB 2: DAFTAR & REKAPITULASI ---
  with tab2:
    st.markdown("#### 📋 Daftar & Rekapitulasi Jurnal Mengajar")
    try:
      data = jurnal_ws.get_all_records()
      if data:
        df = pd.DataFrame(data)
        df.columns = df.columns.str.strip()

        allowed_cols = [
            "Tanggal",
            "Sekolah",
            "Mata_Pelajaran",
            "Kelas",
            "JP_Ke",
            "Topik_Materi",
            "Hadir",
            "Tidak_Hadir",
            "Catatan",
        ]
        existing_cols = [col for col in allowed_cols if col in df.columns]
        df = df[existing_cols]

        if "Sekolah" in df.columns:
          list_sekolah_db = ["-- Semua Sekolah --"] + df[
              "Sekolah"
          ].dropna().unique().tolist()
          pilih_sekolah_filter = st.selectbox(
              "🔍 Filter Berdasarkan Sekolah", options=list_sekolah_db
          )
          if pilih_sekolah_filter != "-- Semua Sekolah --":
            df = df[df["Sekolah"] == pilih_sekolah_filter]

        opsi_filter_kelas = ["-- Semua Kelas --"] + daftar_kelas
        pilih_kelas = st.selectbox(
            "🔍 Filter Berdasarkan Kelas", options=opsi_filter_kelas
        )
        if pilih_kelas != "-- Semua Kelas --":
          df = df[df["Kelas"] == pilih_kelas]

        st.dataframe(df, use_container_width=True, hide_index=True)
      else:
        st.info("Belum ada data Jurnal Mengajar yang tersimpan di database.")
    except Exception as e:
      st.error(f"Gagal memuat data dari Google Sheets: {e}")

  # --- TAB 3: EXPORT LAPORAN DENGAN IDENTITAS & KOLOM SEKOLAH DISEMBUNYIKAN ---
  with tab3:
    st.markdown("#### 📥 Cetak & Export Laporan ke Excel (.xlsx)")
    try:
      data = jurnal_ws.get_all_records()
      if data:
        df_export = pd.DataFrame(data)
        df_export.columns = df_export.columns.str.strip()

        allowed_cols = [
            "Tanggal",
            "Sekolah",
            "Mata_Pelajaran",
            "Kelas",
            "JP_Ke",
            "Topik_Materi",
            "Hadir",
            "Tidak_Hadir",
            "Catatan",
        ]
        existing_cols = [col for col in allowed_cols if col in df_export.columns]
        df_export = df_export[existing_cols]

        col_ex1, col_ex2 = st.columns(2)
        with col_ex1:
          if "Sekolah" in df_export.columns:
            list_sekolah_export = ["-- Semua Sekolah --"] + df_export[
                "Sekolah"
            ].dropna().unique().tolist()
            pilih_sekolah_ex = st.selectbox(
                "🏫 Filter Sekolah untuk Export",
                options=list_sekolah_export,
                key="ex_sekolah",
            )
            if pilih_sekolah_ex != "-- Semua Sekolah --":
              df_export = df_export[df_export["Sekolah"] == pilih_sekolah_ex]

        with col_ex2:
          opsi_filter_kelas_ex = ["-- Semua Kelas --"] + daftar_kelas
          pilih_kelas_ex = st.selectbox(
              "🏫 Filter Kelas untuk Export",
              options=opsi_filter_kelas_ex,
              key="ex_kelas",
          )
          if pilih_kelas_ex != "-- Semua Kelas --":
            df_export = df_export[df_export["Kelas"] == pilih_kelas_ex]

        # Konversi Format Tanggal ke Format Indonesia: Hari, Tanggal-Bulan-Tahun
        days_indo = {
            "Monday": "Senin",
            "Tuesday": "Selasa",
            "Wednesday": "Rabu",
            "Thursday": "Kamis",
            "Friday": "Jumat",
            "Saturday": "Sabtu",
            "Sunday": "Minggu",
        }

        def format_tanggal_indo(val):
          try:
            dt_obj = pd.to_datetime(val)
            nama_hari = days_indo.get(
                dt_obj.strftime("%A"), dt_obj.strftime("%A")
            )
            return f"{nama_hari}, {dt_obj.strftime('%d-%m-%Y')}"
          except:
            return str(val)

        if "Tanggal" in df_export.columns:
          df_export["Tanggal"] = df_export["Tanggal"].apply(
              format_tanggal_indo
          )

        # Sembunyikan kolom 'Sekolah' dari tabel export karena sudah ada di identitas atas
        if "Sekolah" in df_export.columns:
          df_export = df_export.drop(columns=["Sekolah"])

        st.dataframe(df_export, use_container_width=True, hide_index=True)

        # Proses Excel menggunakan openpyxl
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
          df_export.to_excel(writer, index=False, sheet_name="Jurnal Mengajar")

        output.seek(0)
        import openpyxl

        wb = openpyxl.load_workbook(output)
        ws = wb.active

        # Sisipkan 5 baris di bagian atas untuk Judul Identitas
        ws.insert_rows(idx=1, amount=5)

        # Mengisi Header Identitas di Baris 1 sampai 4
        ws["A1"] = "JURNAL MENGAJAR GURU"
        ws["A1"].font = Font(
            name="Calibri", size=14, bold=True, color="1F4E78"
        )
        ws["A2"] = f"Nama Guru : {NAMA_GURU}"
        ws["A2"].font = Font(name="Calibri", size=11, bold=True)
        ws["A3"] = f"Sekolah   : {pilih_sekolah_ex}"
        ws["A3"].font = Font(name="Calibri", size=11, bold=True)
        ws["A4"] = f"Kelas     : {pilih_kelas_ex}"
        ws["A4"].font = Font(name="Calibri", size=11, bold=True)

        # Pengaturan Halaman Siap Cetak (Landscape & Fit to 1 Page Wide)
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        # Header Tabel sekarang berada di baris ke-6 setelah disisipkan
        header_row = 6
        header_fill = PatternFill(
            start_color="1F4E78", end_color="1F4E78", fill_type="solid"
        )
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )

        for col_num in range(1, ws.max_column + 1):
          cell = ws.cell(row=header_row, column=col_num)
          cell.fill = header_fill
          cell.font = header_font
          cell.alignment = Alignment(
              horizontal="center", vertical="center", wrap_text=True
          )
          cell.border = thin_border

        # Styling Data Isi Tabel & Pemberian Border Otomatis (Mulai baris ke-7)
        for row_num in range(header_row + 1, ws.max_row + 1):
          for col_num in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = Font(name="Calibri", size=11)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        # Standar lebar minimum spesifik per kolom agar proporsional dan pas (tanpa kolom Sekolah)
        default_min_widths = {
            "Tanggal": 22,
            "Mata_Pelajaran": 20,
            "Kelas": 12,
            "JP_Ke": 12,
            "Topik_Materi": 30,
            "Hadir": 10,
            "Tidak_Hadir": 14,
            "Catatan": 25,
        }

        for col_num in range(1, ws.max_column + 1):
          col_letter = get_column_letter(col_num)
          col_name = ws.cell(row=header_row, column=col_num).value
          max_len = 0
          for row_num in range(header_row, ws.max_row + 1):
            cell_val = ws.cell(row=row_num, column=col_num).value
            if cell_val is not None:
              max_len = max(max_len, len(str(cell_val)))

          min_w = default_min_widths.get(col_name, 15)
          ws.column_dimensions[col_letter].width = max(max_len + 3, min_w)

        final_output = io.BytesIO()
        wb.save(final_output)
        excel_data = final_output.getvalue()

        nama_file_excel = f"Rekap_Jurnal_{pilih_sekolah_ex.replace(' ', '_')}_{pilih_kelas_ex.replace(' ', '_')}_{date.today()}.xlsx"

        st.download_button(
            label="📥 Download Laporan Excel Siap Cetak (.xlsx)",
            data=excel_data,
            file_name=nama_file_excel,
            mime=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )
      else:
        st.info("Tidak ada data untuk diexport.")
    except Exception as e:
      st.error(f"Gagal menyiapkan data untuk export: {e}")